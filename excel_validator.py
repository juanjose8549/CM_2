import re
from openpyxl import load_workbook
from typing import Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

# Patrones de código malicioso compilados una sola vez (optimización)
MALICIOUS_PATTERNS = [
    re.compile(r'(?i)Auto_?Open'),
    re.compile(r'(?i)Auto_?Close'),
    re.compile(r'(?i)Auto_?Activate'),
    re.compile(r'(?i)Auto_?Deactivate'),
    re.compile(r'(?i)Auto_?Exec'),
    re.compile(r'(?i)Auto_?Run'),
    re.compile(r'(?i)Shell\s*\('),
    re.compile(r'(?i)WScript\.Shell'),
    re.compile(r'(?i)CreateObject\s*\(\s*["\']WScript\.Shell["\']'),
    re.compile(r'(?i)CreateObject\s*\(\s*["\']Shell\.Application["\']'),
    re.compile(r'(?i)Run\s*\('),
    re.compile(r'(?i)Exec\s*\('),
    re.compile(r'(?i)CreateObject\s*\(\s*["\']Scripting\.FileSystemObject["\']'),
    re.compile(r'(?i)FileSystemObject'),
    re.compile(r'(?i)\.DeleteFile\s*\('),
    re.compile(r'(?i)\.DeleteFolder\s*\('),
    re.compile(r'(?i)\.CopyFile\s*\('),
    re.compile(r'(?i)\.MoveFile\s*\('),
    re.compile(r'(?i)\.WriteLine\s*\('),
    re.compile(r'(?i)\.Write\s*\('),
    re.compile(r'(?i)\.SaveToFile\s*\('),
    re.compile(r'(?i)\.OpenTextFile\s*\('),
    re.compile(r'(?i)cmd\s*\.'),
    re.compile(r'(?i)PowerShell\s*\.'),
    re.compile(r'(?i)MSHTA'),
    re.compile(r'(?i)CertUtil'),
    re.compile(r'(?i)BitsAdmin'),
    re.compile(r'(?i)wmic\s+'),
    re.compile(r'(?i)regsvr32\s+'),
    re.compile(r'(?i)URLDownloadToFile'),
    re.compile(r'(?i)WinHttp\.WinHttpRequest'),
    re.compile(r'(?i)XMLHTTP'),
    re.compile(r'(?i)\.send\s*\('),
    re.compile(r'(?i)\.responseText'),
    re.compile(r'(?i)\.responseBody'),
    re.compile(r'(?i)Eval\s*\('),
    re.compile(r'(?i)Execute\b'),
    re.compile(r'(?i)ExecuteGlobal'),
    re.compile(r'(?i)Chr\s*\('),
    re.compile(r'(?i)Asc\s*\('),
    re.compile(r'(?i)ChrW\s*\('),
    re.compile(r'(?i)CryptoMin'),
    re.compile(r'(?i)CoinHive'),
    re.compile(r'(?i)Monero'),
    re.compile(r'(?i)\.mine\s*\('),
    re.compile(r'(?i)CreateObject\s*\(\s*["\']MSXML2\.'),
    re.compile(r'(?i)CreateObject\s*\(\s*["\']WinHttp\.'),
    re.compile(r'(?i)\.open\s*\(\s*["\']GET["\']'),
    re.compile(r'(?i)\.open\s*\(\s*["\']POST["\']'),
    re.compile(r'(?i)VBA\.Shell'),
    re.compile(r'(?i)VBA\.CreateObject'),
    re.compile(r'(?i)VBA\.CallByName'),
    re.compile(r'(?i)\.FormulaHidden\s*=\s*True'),
    re.compile(r'(?i)\.VeryHidden\s*=\s*True'),
]

# Patrón para referencias externas a ejecutables (compilado)
EXTERNAL_EXEC_PATTERN = re.compile(r'(?i)(?:\[.*?\]|\'.*?\'!)(?:.*\.(?:exe|bat|cmd|ps1|vbs|js|jar))')

# Funciones estándar de Excel (optimizado como set de strings en mayúsculas)
STANDARD_EXCEL_FUNCTIONS = {
    # Matemáticas y trigonometría
    'ABS', 'ACOS', 'ACOSH', 'ACOT', 'ACOTH', 'AGGREGATE', 'ARABIC', 'ASIN',
    'ASINH', 'ATAN', 'ATAN2', 'ATANH', 'BASE', 'CEILING', 'CEILING.MATH',
    'CEILING.PRECISE', 'COMBIN', 'COMBINA', 'COS', 'COSH', 'COT', 'COTH',
    'CSC', 'CSCH', 'DECIMAL', 'DEGREES', 'EVEN', 'EXP', 'FACT', 'FACTDOUBLE',
    'FLOOR', 'FLOOR.MATH', 'FLOOR.PRECISE', 'GCD', 'INT', 'ISO.CEILING',
    'LCM', 'LN', 'LOG', 'LOG10', 'LOG2', 'MDETERM', 'MINVERSE', 'MMULT',
    'MOD', 'MROUND', 'MULTINOMIAL', 'MUNIT', 'ODD', 'PI', 'POWER',
    'PRODUCT', 'QUOTIENT', 'RADIANS', 'RAND', 'RANDBETWEEN', 'ROMAN',
    'ROUND', 'ROUNDDOWN', 'ROUNDUP', 'SEC', 'SECH', 'SERIESSUM',
    'SIGN', 'SIN', 'SINH', 'SQRT', 'SQRTPI', 'SUM', 'SUMIF', 'SUMIFS',
    'SUMPRODUCT', 'SUMSQ', 'SUMX2MY2', 'SUMX2PY2', 'SUMXMY2', 'TAN',
    'TANH', 'TRUNC',
    # Estadísticas
    'AVEDEV', 'AVERAGE', 'AVERAGEA', 'AVERAGEIF', 'AVERAGEIFS',
    'BETA.DIST', 'BETA.INV', 'BINOM.DIST', 'BINOM.DIST.RANGE',
    'BINOM.INV', 'CHISQ.DIST', 'CHISQ.DIST.RT', 'CHISQ.INV',
    'CHISQ.INV.RT', 'CHISQ.TEST', 'CONFIDENCE', 'CONFIDENCE.NORM',
    'CONFIDENCE.T', 'CORREL', 'COUNT', 'COUNTA', 'COUNTBLANK',
    'COUNTIF', 'COUNTIFS', 'COVARIANCE', 'COVARIANCE.P', 'COVARIANCE.S',
    'DEVSQ', 'EXPON.DIST', 'F.DIST', 'F.DIST.RT', 'F.INV', 'F.INV.RT',
    'F.TEST', 'FISHER', 'FISHERINV', 'FORECAST', 'FORECAST.ETS',
    'FORECAST.ETS.CONFINT', 'FORECAST.ETS.SEASONALITY',
    'FORECAST.ETS.STAT', 'FORECAST.LINEAR', 'FREQUENCY', 'GAMMA',
    'GAMMA.DIST', 'GAMMA.INV', 'GAMMALN', 'GAMMALN.PRECISE', 'GAUSS',
    'GEOMEAN', 'GROWTH', 'HARMEAN', 'HYPGEOM.DIST', 'INTERCEPT',
    'KURT', 'LARGE', 'LINEST', 'LOGEST', 'LOGNORM.DIST', 'LOGNORM.INV',
    'MAX', 'MAXA', 'MAXIFS', 'MEDIAN', 'MIN', 'MINA', 'MINIFS', 'MODE',
    'MODE.MULT', 'MODE.SNGL', 'NEGBINOM.DIST', 'NORM.DIST', 'NORM.INV',
    'NORM.S.DIST', 'NORM.S.INV', 'PEARSON', 'PERCENTILE', 'PERCENTILE.EXC',
    'PERCENTILE.INC', 'PERCENTRANK', 'PERCENTRANK.EXC', 'PERCENTRANK.INC',
    'PERMUT', 'PERMUTATIONA', 'PHI', 'POISSON.DIST', 'PROB', 'QUARTILE',
    'QUARTILE.EXC', 'QUARTILE.INC', 'RANK', 'RANK.AVG', 'RANK.EQ',
    'RSQ', 'SKEW', 'SKEW.P', 'SLOPE', 'SMALL', 'STANDARDIZE', 'STDEV',
    'STDEV.P', 'STDEV.S', 'STDEVA', 'STDEVPA', 'STEYX', 'T.DIST',
    'T.DIST.2T', 'T.DIST.RT', 'T.INV', 'T.INV.2T', 'T.TEST', 'TREND',
    'TRIMMEAN', 'VAR', 'VAR.P', 'VAR.S', 'VARA', 'VARPA', 'WEIBULL.DIST',
    'Z.TEST',
    # Texto
    'ARRAYTOTEXT', 'ASC', 'BAHTTEXT', 'CHAR', 'CLEAN', 'CODE', 'CONCAT',
    'CONCATENATE', 'DBCS', 'DOLLAR', 'EXACT', 'FIND', 'FINDB', 'FIXED',
    'LEFT', 'LEFTB', 'LEN', 'LENB', 'LOWER', 'MID', 'MIDB', 'NUMBERVALUE',
    'PHONETIC', 'PROPER', 'REPLACE', 'REPLACEB', 'REPT', 'RIGHT', 'RIGHTB',
    'SEARCH', 'SEARCHB', 'SUBSTITUTE', 'T', 'TEXT', 'TEXTJOIN',
    'TRIM', 'UNICHAR', 'UNICODE', 'UPPER', 'VALUE', 'VALUETOTEXT',
    # Fecha y hora
    'DATE', 'DATEVALUE', 'DAY', 'DAYS', 'DAYS360', 'EDATE', 'EOMONTH',
    'HOUR', 'ISOWEEKNUM', 'MINUTE', 'MONTH', 'NETWORKDAYS',
    'NETWORKDAYS.INTL', 'NOW', 'SECOND', 'TIME', 'TIMEVALUE', 'TODAY',
    'WEEKDAY', 'WEEKNUM', 'WORKDAY', 'WORKDAY.INTL', 'YEAR', 'YEARFRAC',
    # Búsqueda y referencia
    'ADDRESS', 'AREAS', 'CHOOSE', 'CHOOSECOLS', 'CHOOSEROWS', 'COLUMN',
    'COLUMNS', 'DROP', 'EXPAND', 'FILTER', 'FORMULATEXT', 'GETPIVOTDATA',
    'HLOOKUP', 'HSTACK', 'HYPERLINK', 'INDEX', 'INDIRECT', 'LOOKUP',
    'MATCH', 'OFFSET', 'ROW', 'ROWS', 'RANDARRAY', 'SORT', 'SORTBY',
    'SEQUENCE', 'SINGLE', 'TAKE', 'TOCOL', 'TOROW', 'TRANSPOSE',
    'UNIQUE', 'VLOOKUP', 'VSTACK', 'WRAPCOLS', 'WRAPROWS', 'XLOOKUP',
    'XMATCH',
    # Lógica
    'AND', 'BYCOL', 'BYROW', 'FALSE', 'IF', 'IFERROR', 'IFNA', 'IFS',
    'LAMBDA', 'LET', 'MAKEARRAY', 'MAP', 'NOT', 'OR', 'REDUCE', 'SCAN',
    'SWITCH', 'TRUE', 'XOR',
    # Información
    'CELL', 'ERROR.TYPE', 'INFO', 'ISBLANK', 'ISERR', 'ISERROR',
    'ISEVEN', 'ISFORMULA', 'ISLOGICAL', 'ISNA', 'ISNONTEXT', 'ISNUMBER',
    'ISODD', 'ISOMITTED', 'ISREF', 'ISTEXT', 'N', 'NA', 'SHEET',
    'SHEETS', 'TYPE',
    # Finanzas
    'ACCRINT', 'ACCRINTM', 'AMORDEGRC', 'AMORLINC', 'COUPDAYBS',
    'COUPDAYS', 'COUPDAYSNC', 'COUPNCD', 'COUPNUM', 'COUPPCD',
    'CUMIPMT', 'CUMPRINC', 'DB', 'DDB', 'DISC', 'DOLLARDE', 'DOLLARFR',
    'DURATION', 'EFFECT', 'FV', 'FVSCHEDULE', 'INTRATE', 'IPMT',
    'IRR', 'ISPMT', 'MDURATION', 'MIRR', 'NOMINAL', 'NPER', 'NPV',
    'ODDFPRICE', 'ODDFYIELD', 'ODDLPRICE', 'ODDLYIELD', 'PDURATION',
    'PMT', 'PPMT', 'PRICE', 'PRICEDISC', 'PRICEMAT', 'PV', 'RATE',
    'RECEIVED', 'RRI', 'SLN', 'STOCKHISTORY', 'SYD', 'TBILLEQ',
    'TBILLPRICE', 'TBILLYIELD', 'VDB', 'XIRR', 'XNPV', 'YIELD',
    'YIELDDISC', 'YIELDMAT',
    # Ingeniería
    'BESSELI', 'BESSELJ', 'BESSELK', 'BESSELY', 'BIN2DEC', 'BIN2HEX',
    'BIN2OCT', 'BITAND', 'BITLSHIFT', 'BITOR', 'BITRSHIFT', 'BITXOR',
    'COMPLEX', 'CONVERT', 'DEC2BIN', 'DEC2HEX', 'DEC2OCT', 'DELTA',
    'ERF', 'ERF.PRECISE', 'ERFC', 'ERFC.PRECISE', 'GESTEP', 'HEX2BIN',
    'HEX2DEC', 'HEX2OCT', 'IMABS', 'IMAGINARY', 'IMARGUMENT',
    'IMCONJUGATE', 'IMCOS', 'IMCOSH', 'IMCOT', 'IMCSC', 'IMCSCH',
    'IMDIV', 'IMEXP', 'IMLN', 'IMLOG10', 'IMLOG2', 'IMPOWER', 'IMPRODUCT',
    'IMREAL', 'IMSEC', 'IMSECH', 'IMSIN', 'IMSINH', 'IMSQRT', 'IMSUB',
    'IMSUM', 'IMTAN', 'OCT2BIN', 'OCT2DEC', 'OCT2HEX',
    # Base de datos
    'DAVERAGE', 'DCOUNT', 'DCOUNTA', 'DGET', 'DMAX', 'DMIN', 'DPRODUCT',
    'DSTDEV', 'DSTDEVP', 'DSUM', 'DVAR', 'DVARP',
    # Cubo
    'CUBESET', 'CUBESETCOUNT', 'CUBEVALUE',
    # Web
    'ENCODEURL', 'WEBSERVICE', 'FILTERXML',
    # Compatibilidad
    'R1C1', 'A1',
    'XLL', 'UDF',
}

# Palabras sospechosas para contenido de celdas (compilado como set)
SUSPICIOUS_WORDS = {
    'auto_open', 'auto_close', 'auto_activate',
    'shell', 'exec', 'run', 'cmd', 'powershell',
    'wscript', 'createobject', 'filesystemobject',
}


def _analyze_formula(formula: str, sheet_name: str, cell_coordinate: str) -> Optional[Dict]:
    """
    Analiza una fórmula en busca de patrones maliciosos.
    Retorna dict con info de la amenaza o None si es segura.
    """
    formula_upper = formula.upper()
    
    # Verificar si contiene funciones estándar de Excel (rápido)
    # Dividir en tokens por paréntesis, espacios, operadores
    formula_upper = formula.replace('=', '', 1) if formula.startswith('=') else formula
    # Extraer posibles nombres de funciones
    tokens = set(formula_upper.replace('(', ' ').replace(')', ' ').replace(',', ' ').replace(';', ' ').split())
    
    # Buscar patrones maliciosos
    for compiled_pattern in MALICIOUS_PATTERNS:
        match = compiled_pattern.search(formula)
        if match:
            # Verificar que no sea una función estándar
            is_standard = bool(tokens & STANDARD_EXCEL_FUNCTIONS)
            if not is_standard:
                return {
                    "sheet": sheet_name,
                    "cell": cell_coordinate,
                    "pattern_matched": compiled_pattern.pattern,
                    "formula": formula[:200],
                }
    
    return None


def _analyze_sheet(sheet_ws, sheet_name: str) -> Dict:
    """
    Analiza una hoja de Excel de forma rápida.
    Retorna dict con resultados del análisis de la hoja.
    """
    sheet_result = {
        "cells_analyzed": 0,
        "formulas_found": 0,
        "malicious_patterns_found": [],
        "unusual_cells": [],
        "errors": [],
    }
    
    for row in sheet_ws.iter_rows():
        for cell in row:
            sheet_result["cells_analyzed"] += 1
            cell_value = cell.value
            
            if cell_value is None:
                continue
            
            # Analizar celdas con fórmula (comienza con '=')
            if isinstance(cell_value, str):
                if cell_value.startswith('='):
                    sheet_result["formulas_found"] += 1
                    threat = _analyze_formula(cell_value, sheet_name, cell.coordinate)
                    if threat:
                        sheet_result["malicious_patterns_found"].append(threat)
                        sheet_result["errors"].append(
                            f"Celda {cell.coordinate} en hoja '{sheet_name}': "
                            f"posible código malicioso detectado"
                        )
                
                # Verificar palabras sospechosas en celdas de texto
                cell_lower = cell_value.lower()
                for word in SUSPICIOUS_WORDS:
                    if word in cell_lower:
                        sheet_result["unusual_cells"].append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "suspicious_content": word,
                        })
                
                # Verificar referencias externas a ejecutables
                if EXTERNAL_EXEC_PATTERN.search(cell_value):
                    sheet_result["errors"].append(
                        f"Celda {cell.coordinate} en hoja '{sheet_name}': "
                        f"referencia externa a archivo ejecutable detectada"
                    )
    
    return sheet_result


def validate_excel_safety(file_path: str) -> Dict[str, Any]:
    """
    Valida que un archivo Excel no contenga código malicioso.
    Optimizado para archivos grandes usando análisis por lotes.
    
    Args:
        file_path: Ruta al archivo Excel
    
    Returns:
        Dict con resultado de la validación
    """
    result = {
        "safe": True,
        "warnings": [],
        "errors": [],
        "details": {
            "sheets": 0,
            "cells_analyzed": 0,
            "formulas_found": 0,
            "malicious_patterns_found": [],
            "unusual_cells": [],
        }
    }
    
    try:
        workbook = load_workbook(file_path, data_only=False, keep_vba=True, read_only=True)
        result["details"]["sheets"] = len(workbook.sheetnames)
        
        # Verificar VBA (optimizado)
        try:
            vba_project = (
                getattr(workbook, 'vbaProject', None) or 
                getattr(workbook, 'vba_project', None)
            )
            if vba_project is not None:
                result["warnings"].append(
                    "El archivo contiene macros VBA. Se requiere revisión manual."
                )
                result["details"]["has_vba"] = True
        except Exception:
            pass
        
        # Analizar cada hoja
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_result = _analyze_sheet(worksheet, sheet_name)
            
            # Acumular resultados
            result["details"]["cells_analyzed"] += sheet_result["cells_analyzed"]
            result["details"]["formulas_found"] += sheet_result["formulas_found"]
            result["details"]["malicious_patterns_found"].extend(
                sheet_result["malicious_patterns_found"]
            )
            result["details"]["unusual_cells"].extend(sheet_result["unusual_cells"])
            
            if sheet_result["malicious_patterns_found"]:
                result["safe"] = False
            
            result["errors"].extend(sheet_result["errors"])
        
        workbook.close()
        
    except Exception as e:
        result["safe"] = False
        result["errors"].append(f"Error al procesar el archivo: {str(e)}")
    
    return result


def get_safe_excel_content(file_path: str) -> Dict[str, Any]:
    """
    Lee el contenido seguro de un archivo Excel (solo valores, sin fórmulas).
    Optimizado para archivos grandes usando read_only mode.
    
    Args:
        file_path: Ruta al archivo Excel
    
    Returns:
        Dict con el contenido del archivo
    """
    content = {
        "sheets": {},
        "metadata": {
            "filename": file_path.split('/')[-1] if '/' in file_path else file_path,
        }
    }
    
    try:
        # Usar read_only=True y data_only=True para máxima eficiencia
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        max_columns = 0
        total_rows = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_data = []
            
            for row in worksheet.iter_rows(values_only=True):
                cleaned_row = ["" if cell_value is None else str(cell_value) for cell_value in row]
                sheet_data.append(cleaned_row)
                max_columns = max(max_columns, len(cleaned_row))
                total_rows += 1
            
            content["sheets"][sheet_name] = sheet_data
        
        content["metadata"]["columns"] = max_columns
        content["metadata"]["rows"] = total_rows
        
        workbook.close()
        
    except Exception as e:
        content["error"] = str(e)
    
    return content
